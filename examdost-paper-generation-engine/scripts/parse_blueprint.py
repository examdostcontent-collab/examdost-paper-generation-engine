"""
parse_blueprint.py — read Skill-01's Exam Master Blueprint .xlsx into JSON.

Usage:
    python parse_blueprint.py <blueprint.xlsx> <out_blueprint.json>

Skill-01 (exam-psychometrician) emits a 4-sheet workbook:
  Sheet "Weightage Matrix"  (Part A)
  Sheet "Exam DNA Profile"  (Part B — Layer A macro + Layer B micro)
  Sheet "Structural Template" (Part C)        [optional]
  Sheet "Generation Spec"   (Part D — Archetype Bank)  [optional]

This reads them into one JSON with best-effort structured fields PLUS a
guaranteed `raw_text` dump of every sheet (so even if a sheet's layout drifts,
no information is lost — the generator can read the raw grid). Sheet names are
matched case/spacing-tolerantly.

Dependencies: openpyxl.
"""
from __future__ import annotations

import json
import sys

from openpyxl import load_workbook


def _norm(s):
    return str(s or "").strip().lower()


def _cells(row):
    """Row of values -> list of trimmed strings (None -> '')."""
    return [("" if v is None else str(v).strip()) for v in row]


def _nonempty(cells):
    return [c for c in cells if c]


def _find_sheet(wb, *needles):
    """Return the first worksheet whose title contains any needle (lowercased)."""
    for ws in wb.worksheets:
        t = _norm(ws.title)
        if any(n in t for n in needles):
            return ws
    return None


def _grid(ws):
    return [_cells(r) for r in ws.iter_rows(values_only=True)]


def _raw_text(ws):
    lines = []
    for cells in _grid(ws):
        if _nonempty(cells):
            lines.append("\t".join(cells).rstrip())
    return "\n".join(lines)


# ----------------------------- Part A -------------------------------------
def parse_weightage(ws):
    rows = _grid(ws)
    # locate the header row: contains "subject" and "questions"
    hdr_idx = None
    for i, cells in enumerate(rows):
        low = [_norm(c) for c in cells]
        if "subject" in low and any("question" in c for c in low):
            hdr_idx = i
            break
    if hdr_idx is None:
        return []
    header = [_norm(c) for c in rows[hdr_idx]]

    def col(*names):
        for j, h in enumerate(header):
            if any(n in h for n in names):
                return j
        return None

    ci = {
        "subject": col("subject"),
        "chapter": col("chapter", "topic"),
        "questions": col("question"),
        "marks": col("mark"),
        "weight_pct": col("weight"),
        "difficulty": col("difficult"),
        "format_mix": col("format"),
        "notes": col("note"),
    }
    out = []
    for cells in rows[hdr_idx + 1:]:
        if not _nonempty(cells):
            continue
        first = _norm(cells[0]) if cells else ""
        if first in ("total", "totals"):
            break
        rec = {}
        for key, j in ci.items():
            rec[key] = cells[j] if (j is not None and j < len(cells)) else ""
        if not rec.get("subject") and not rec.get("chapter"):
            continue
        out.append(rec)
    return out


# ----------------------------- Part B -------------------------------------
MICRO_KEYS = ["trap_logic", "variable_types", "calculation_steps", "archetypes"]


def parse_dna(ws):
    rows = _grid(ws)
    macro, micro = {}, {}
    mode = None          # "macro" | "micro"
    micro_cols = None    # column index map for the micro table
    for cells in rows:
        ne = _nonempty(cells)
        if not ne:
            continue
        c0 = cells[0]
        low0 = _norm(c0)
        # band detection
        if "layer a" in low0 or ("macro" in low0 and "layer" in low0):
            mode = "macro"
            continue
        if "layer b" in low0 or ("micro" in low0 and "layer" in low0):
            mode = "micro"
            micro_cols = None
            continue
        if mode == "macro":
            # Aspect | Detail (detail lives in col 2 after a merge)
            detail = ""
            for v in cells[1:]:
                if v:
                    detail = v
                    break
            if c0 and detail:
                macro[c0] = detail
        elif mode == "micro":
            low = [_norm(c) for c in cells]
            if micro_cols is None:
                # this is the header row (subject + the micro keys)
                micro_cols = {}
                for j, h in enumerate(low):
                    if "subject" in h:
                        micro_cols["subject"] = j
                    elif "trap" in h:
                        micro_cols["trap_logic"] = j
                    elif "variable" in h:
                        micro_cols["variable_types"] = j
                    elif "calc" in h or "step" in h:
                        micro_cols["calculation_steps"] = j
                    elif "archetype" in h:
                        micro_cols["archetypes"] = j
                continue
            sj = micro_cols.get("subject", 0)
            subj = cells[sj] if sj < len(cells) else ""
            if not subj:
                continue
            micro[subj] = {
                k: (cells[micro_cols[k]] if k in micro_cols and micro_cols[k] < len(cells) else "")
                for k in MICRO_KEYS
            }
    dna = {}
    if macro:
        dna["macro"] = macro
    if micro:
        dna["micro"] = micro
    return dna


# ----------------------------- Part C -------------------------------------
def _parse_table_after(rows, start_idx):
    """Read a header row + data rows starting at start_idx until a blank/band row.
    Returns (list_of_dicts, next_idx)."""
    i = start_idx
    while i < len(rows) and not _nonempty(rows[i]):
        i += 1
    if i >= len(rows):
        return [], i
    header = [h for h in rows[i]]
    keys = [_norm(h).replace(" ", "_") for h in header]
    i += 1
    out = []
    while i < len(rows):
        cells = rows[i]
        if not _nonempty(cells):
            break
        # a single-cell row is likely the next band -> stop
        if len(_nonempty(cells)) == 1 and not cells[0][:1].isdigit() and len(header) > 1:
            # could be a band; stop unless it actually fills only first col of a 1-col table
            break
        rec = {}
        for j, k in enumerate(keys):
            if k:
                rec[k] = cells[j] if j < len(cells) else ""
        out.append(rec)
        i += 1
    return out, i


def parse_structural(ws):
    rows = _grid(ws)
    st = {"sections": [], "question_matrix": [], "format_matrix": [], "hard_constraints": []}
    i = 0
    n = len(rows)
    while i < n:
        ne = _nonempty(rows[i])
        if not ne:
            i += 1
            continue
        band = _norm(rows[i][0])
        if "section boundaries" in band:
            st["sections"], i = _parse_table_after(rows, i + 1)
            continue
        if "question matrix" in band:
            st["question_matrix"], i = _parse_table_after(rows, i + 1)
            continue
        if "format matrix" in band:
            st["format_matrix"], i = _parse_table_after(rows, i + 1)
            continue
        if "hard constraint" in band:
            i += 1
            while i < n:
                cells = rows[i]
                if not _nonempty(cells):
                    i += 1
                    continue
                txt = cells[0].lstrip("•").strip()
                # stop if we hit another recognisable band (unlikely after constraints)
                st["hard_constraints"].append(txt)
                i += 1
            break
        i += 1
    # drop empties
    if not any(st.values()):
        return {}
    return {k: v for k, v in st.items() if v}


# ----------------------------- Part D -------------------------------------
GEN_FIELDS = ["id", "subject", "chapter", "topic", "format", "difficulty",
              "expected_per_paper", "template", "solution_path",
              "distractor_rules", "numeric_profile", "phrasing_pattern", "seed_examples"]


def parse_generation(ws):
    rows = _grid(ws)
    hdr_idx = None
    for i, cells in enumerate(rows):
        low = [_norm(c) for c in cells]
        if any("template" in c for c in low) and any("distractor" in c for c in low):
            hdr_idx = i
            break
    if hdr_idx is None:
        return {}
    header = [_norm(c) for c in rows[hdr_idx]]

    def col(*names):
        for j, h in enumerate(header):
            if any(n in h for n in names):
                return j
        return None

    cmap = {
        "id": col("id"), "subject": col("subject"), "chapter": col("chapter"),
        "topic": col("topic"), "format": col("format"), "difficulty": col("diff"),
        "expected_per_paper": col("exp"), "template": col("template"),
        "solution_path": col("solution"), "distractor_rules": col("distractor"),
        "numeric_profile": col("numeric", "parameter"), "phrasing_pattern": col("phrasing"),
        "seed_examples": col("seed"),
    }
    arch = []
    for cells in rows[hdr_idx + 1:]:
        if not _nonempty(cells):
            continue
        rec = {}
        for k, j in cmap.items():
            val = cells[j] if (j is not None and j < len(cells)) else ""
            if k in ("distractor_rules", "seed_examples") and val:
                parts = [p.strip(" •—") for p in val.replace("— — —", "\n").splitlines()]
                rec[k] = [p for p in parts if p]
            else:
                rec[k] = val
        if rec.get("id") or rec.get("template"):
            arch.append(rec)
    return {"archetypes": arch} if arch else {}


# ----------------------------- driver -------------------------------------
def parse(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    out = {"source": path, "sheets_found": list(wb.sheetnames), "raw_text": {}}

    ws_a = _find_sheet(wb, "weightage", "weight matrix", "part a")
    ws_b = _find_sheet(wb, "dna", "exam dna", "part b")
    ws_c = _find_sheet(wb, "structural", "part c")
    ws_d = _find_sheet(wb, "generation", "archetype", "part d")

    if ws_a:
        out["weightage_matrix"] = parse_weightage(ws_a)
        out["raw_text"][ws_a.title] = _raw_text(ws_a)
    if ws_b:
        out["exam_dna"] = parse_dna(ws_b)
        out["raw_text"][ws_b.title] = _raw_text(ws_b)
    if ws_c:
        st = parse_structural(ws_c)
        if st:
            out["structural_template"] = st
        out["raw_text"][ws_c.title] = _raw_text(ws_c)
    if ws_d:
        gs = parse_generation(ws_d)
        if gs:
            out["generation_spec"] = gs
        out["raw_text"][ws_d.title] = _raw_text(ws_d)

    # any sheet not matched above -> still dump raw so nothing is lost
    matched = {ws.title for ws in (ws_a, ws_b, ws_c, ws_d) if ws}
    for ws in wb.worksheets:
        if ws.title not in matched and ws.title not in out["raw_text"]:
            out["raw_text"][ws.title] = _raw_text(ws)
    wb.close()
    return out


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    data = parse(sys.argv[1])
    with open(sys.argv[2], "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    # coverage summary
    print(f"Wrote {sys.argv[2]}")
    print(f"  Sheets found: {', '.join(data['sheets_found'])}")
    wm = data.get("weightage_matrix") or []
    dna = data.get("exam_dna") or {}
    st = data.get("structural_template") or {}
    gs = (data.get("generation_spec") or {}).get("archetypes") or []
    print(f"  Part A Weightage rows : {len(wm)}")
    print(f"  Part B macro aspects  : {len(dna.get('macro', {}))}   micro subjects: {len(dna.get('micro', {}))}")
    print(f"  Part C structural     : {'yes' if st else 'MISSING'}"
          + (f" (sections={len(st.get('sections', []))}, constraints={len(st.get('hard_constraints', []))})" if st else ""))
    print(f"  Part D archetypes     : {len(gs)}")
    missing = []
    if not wm:
        missing.append("Part A Weightage Matrix")
    if not dna:
        missing.append("Part B Exam DNA")
    if not st:
        missing.append("Part C Structural Template")
    if not gs:
        missing.append("Part D Generation Spec")
    if missing:
        print("  ! WARNING — could not extract: " + "; ".join(missing)
              + ". Check raw_text in the JSON, or ask the user for that part.")


if __name__ == "__main__":
    main()
