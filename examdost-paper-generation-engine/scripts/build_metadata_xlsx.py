"""
build_metadata_xlsx.py — emit the per-question classification spreadsheet
(the "Mock-01 Metadata"-style sheet: one row per question, 7 fixed columns).

Usage:
    python build_metadata_xlsx.py <questions.json> <out.xlsx> [--taxonomy <master.xlsx>]

Input JSON — any of these shapes is accepted (use whichever is convenient):
    1. A bare list of question dicts:                       [ {...}, {...} ]
    2. {"set_name": "...", "questions": [ {...}, ... ]}     (preferred)
    3. paper-generator's paper.json: {"sections":[{"questions":[...]}]}

Each question dict carries the fields the classifier worked out:
    {
      "subject":    "Network Theory",      # canonical subject name (master sheet)
      "chapter":    "Network Theorems",    # canonical chapter name
      "topic":      "Maximum Power Transfer",  # canonical type/topic name
      "difficulty": 2,                      # 1 Easy / 2 Moderate / 3 Difficult
      "marks":      1,
      "type":       "Numerical MCQ"         # one of the 7 canonical type labels,
                                            #   OR "MCQ" (+ optional "type_detail")
    }

Columns (exactly the agreed template, in this order):
    Question number | Subject Name | Chapter number | type number |
    difficulty index | Marks | TheoryMCQ/Numerical MCQ/NAT/MSQ/StatementType/Assertion-Reason

- Question number  : running 1..N in the order the questions appear.
- Chapter number   : the chapter's number in the master taxonomy for that subject.
- type number      : the Type No of the question's topic within that chapter.
- difficulty index : 1 / 2 / 3 from q.difficulty (falls back to an inference).
- last column      : the canonical question type.

Chapter/type numbers come from the canonical taxonomy (load_taxonomy). When a
question's subject/chapter/topic can't be matched, those cells are left blank and
the row is reported so the user can fix the tag or extend the master.

Dependencies: openpyxl, load_taxonomy (sibling).
"""
from __future__ import annotations

import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["Question number", "Subject Name", "Chapter number", "type number",
           "difficulty index", "Marks",
           "TheoryMCQ/Numerical MCQ/NAT/MSQ/StatementType/Assertion-Reason"]

# canonical final labels the last column may hold
CANON_TYPES = {
    "theorymcq": "TheoryMCQ",
    "theory mcq": "TheoryMCQ",
    "numerical mcq": "Numerical MCQ",
    "numericalmcq": "Numerical MCQ",
    "nat": "NAT",
    "msq": "MSQ",
    "statementtype": "StatementType",
    "statement type": "StatementType",
    "statement": "StatementType",
    "assertion-reason": "Assertion-Reason",
    "assertion reason": "Assertion-Reason",
    "ar": "Assertion-Reason",
}


def _norm(s):
    return str(s or "").strip().lower()


def _put(d, k, v):
    """Aggregate across the split chapter objects load_taxonomy emits (one per
    topic row, with the chapter/type number only on the first row). Keep the
    first non-blank value; never let a trailing blank clobber a real number."""
    if v not in (None, ""):
        if not d.get(k):
            d[k] = v
    else:
        d.setdefault(k, v)


def build_lookup(tax):
    chap, typ = {}, {}
    for subj, sd in (tax.get("subjects") or {}).items():
        sl = _norm(subj)
        for c in sd.get("chapters", []):
            cl = _norm(c.get("name"))
            _put(chap, (sl, cl), c.get("no"))
            for tp in c.get("topics") or []:
                _put(typ, (sl, cl, _norm(tp.get("name"))), tp.get("no"))
    return chap, typ


def excel_type(q):
    """Resolve the canonical type label. The classifier normally writes the final
    label straight into q['type']; we still handle a raw 'MCQ' (+ optional
    type_detail / solution shape) so a paper.json can be fed in unchanged."""
    raw = _norm(q.get("type"))
    if raw in CANON_TYPES:
        return CANON_TYPES[raw]
    if raw.upper() == "MCQ":
        detail = _norm(q.get("type_detail"))
        if "numer" in detail:
            return "Numerical MCQ"
        if "theory" in detail:
            return "TheoryMCQ"
        sol = q.get("solution") or {}
        if sol.get("calculation"):
            return "Numerical MCQ"
        if sol.get("option_analysis"):
            return "TheoryMCQ"
        return "TheoryMCQ"
    # unknown — hand back whatever was given so it's visible, not silently dropped
    return q.get("type", "")


def difficulty(q):
    d = q.get("difficulty")
    if str(d) in ("1", "2", "3"):
        return int(d)
    # tolerant text fallbacks
    dl = _norm(d)
    if dl in ("easy", "e"):
        return 1
    if dl in ("moderate", "medium", "mod", "m"):
        return 2
    if dl in ("difficult", "hard", "d", "h"):
        return 3
    # last-resort inference from solution shape (paper.json case)
    sol = q.get("solution") or {}
    if sol.get("calculation"):
        return 3 if len(sol["calculation"]) > 3 else 2
    if _norm(q.get("type")).startswith("statement"):
        return 2
    return 1


def iter_q(data):
    """Yield question dicts from any of the accepted input shapes."""
    if isinstance(data, list):
        yield from data
        return
    if isinstance(data, dict):
        if data.get("questions"):
            yield from data["questions"]
            return
        if data.get("sections"):
            for sec in data["sections"]:
                yield from sec.get("questions", [])
            return
    # nothing recognized
    return


def build(in_path, out_path, tax_path=None):
    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)
    questions = list(iter_q(data))
    if not questions:
        print("  ! no questions found in the input JSON — check the shape "
              "(expected a list, {'questions':[...]}, or {'sections':[...]})")

    # taxonomy
    here = os.path.dirname(os.path.abspath(__file__))
    sys.path.insert(0, here)
    import load_taxonomy as lt
    tax_path = tax_path or os.path.join(here, "..", "assets", "chapter_topic_master.xlsx")
    chap, typ = ({}, {})
    if os.path.exists(tax_path):
        chap, typ = build_lookup(lt.parse(tax_path))
    else:
        print(f"  ! taxonomy not found at {tax_path} — chapter/type numbers left blank")

    wb = Workbook()
    ws = wb.active
    ws.title = "Question Metadata"
    hdr_fill = PatternFill("solid", fgColor="1F2A44")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D8DCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for j, h in enumerate(HEADERS, start=1):
        c = ws.cell(1, j, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    unmapped = []
    diff_spread = {1: 0, 2: 0, 3: 0}
    r = 2
    for i, q in enumerate(questions, start=1):
        # Use the source's own question number when present, so the sheet lines up
        # with the original paper; otherwise fall back to a running 1..N.
        n = (q.get("number") or q.get("q_no") or q.get("question_number")
             or q.get("_source_q") or i)
        subj = q.get("subject", "")
        sl, cl = _norm(subj), _norm(q.get("chapter"))
        cno = chap.get((sl, cl)) or None       # treat "" (blank in master) as missing
        tno = typ.get((sl, cl, _norm(q.get("topic")))) or None
        if cno is None or tno is None:
            unmapped.append((n, subj, q.get("chapter"), q.get("topic"),
                             cno is not None, tno is not None))
        d = difficulty(q)
        diff_spread[d] = diff_spread.get(d, 0) + 1
        vals = [n, subj, cno if cno is not None else "", tno if tno is not None else "",
                d, q.get("marks", ""), excel_type(q)]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(r, j, v)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if j != 2 else "left",
                                       vertical="center")
        r += 1

    widths = [16, 24, 14, 12, 14, 8, 52]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    wb.save(out_path)

    nq = len(questions)
    print(f"Wrote {out_path}  ({nq} question rows)")
    print(f"  difficulty spread — 1 Easy: {diff_spread.get(1,0)} | "
          f"2 Moderate: {diff_spread.get(2,0)} | 3 Difficult: {diff_spread.get(3,0)}")
    if unmapped:
        print(f"  ! {len(unmapped)} row(s) had no full taxonomy match "
              f"(chapter and/or type number left blank):")
        for n, s, c, t, has_c, has_t in unmapped[:20]:
            miss = []
            if not has_c:
                miss.append("chapter")
            if not has_t:
                miss.append("type")
            print(f"      Q{n}: {s} > {c} > {t}   (no {', '.join(miss)} match)")
        if len(unmapped) > 20:
            print(f"      ... and {len(unmapped) - 20} more")
        print("  -> fix the subject/chapter/topic names to match the master, "
              "or add the missing entry to the master taxonomy.")


def main():
    args = list(sys.argv[1:])
    tax = None
    if "--taxonomy" in args:
        i = args.index("--taxonomy")
        tax = args[i + 1]
        del args[i:i + 2]
    if len(args) != 2:
        print(__doc__)
        sys.exit(1)
    build(args[0], args[1], tax)


if __name__ == "__main__":
    main()
