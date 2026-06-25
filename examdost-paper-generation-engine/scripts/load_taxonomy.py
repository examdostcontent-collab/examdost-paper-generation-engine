"""
load_taxonomy.py — parse the ExamDost Chapter & Topic master sheet into a clean
canonical taxonomy, so every paper analysis uses identical subject / chapter /
topic naming across the whole library.

Usage:
    python load_taxonomy.py [path-to-master.xlsx] [--flat] [--subject "Name"]

Defaults to the bundled copy at assets/chapter_topic_master.xlsx. Pass a path to
read a newer master (e.g. the user's updated sheet in Downloads).

Output (default): JSON tree
    { "subjects": { "<Subject>": { "chapters": [
        { "no": "1", "name": "<Chapter>", "note": "<applicability>",
          "topics": [ { "no": "1", "name": "<Type/Topic>", "note": "" } ] } ] } } }

`--flat` prints one canonical "Subject > Chapter > Topic" path per line — handy
for a quick eyeball or for pasting a vocabulary list.

The master is hand-maintained, so this parser is defensive: it locates the
header row per sheet, forward-fills merged chapter cells, tolerates sheets that
have chapters but no topics, captures a trailing applicability/marks note, and
handles multi-table sheets (several side-by-side chapter tables on one sheet) by
segmenting the header row on blank columns.

Dependencies: openpyxl.
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT = Path(__file__).resolve().parent.parent / "assets" / "chapter_topic_master.xlsx"

CHAP_NO = re.compile(r"chapter\s*(no\.?|number)", re.I)
CHAP_NAME = re.compile(r"chapter\s*name", re.I)
TYPE_NO = re.compile(r"type\s*no", re.I)
TYPE_NAME = re.compile(r"type\s*name", re.I)


def _norm(v) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _find_header_row(rows):
    """Return index of the first row that looks like a column header."""
    for i, row in enumerate(rows):
        cells = [_norm(c) for c in row]
        if any(CHAP_NAME.search(c) for c in cells) and any(
            CHAP_NO.search(c) or c.lower() in ("chapter no.", "chapter number") for c in cells
        ):
            return i
    # fallback: any row mentioning "Chapter Name"
    for i, row in enumerate(rows):
        if any(CHAP_NAME.search(_norm(c)) for c in row):
            return i
    return None


def _segments(header_cells):
    """Split header columns into contiguous non-blank segments (blocks)."""
    segs = []
    cur = []
    for j, c in enumerate(header_cells):
        if c:
            cur.append(j)
        else:
            if cur:
                segs.append(cur)
                cur = []
    if cur:
        segs.append(cur)
    return segs


def _block_cols(header_cells, seg):
    """Within a segment, identify the chapter/type column indices."""
    cn = cnm = tn = tnm = None
    for j in seg:
        c = header_cells[j]
        if cn is None and CHAP_NO.search(c):
            cn = j
        elif cnm is None and CHAP_NAME.search(c):
            cnm = j
        elif tn is None and TYPE_NO.search(c):
            tn = j
        elif tnm is None and TYPE_NAME.search(c):
            tnm = j
    if cnm is None:
        return None
    return {"cn": cn, "cnm": cnm, "tn": tn, "tnm": tnm}


def _parse_block(rows, hdr_idx, cols, note_range, super_label=""):
    """Walk data rows under the header, forward-filling chapters.

    note_range is (start, end) of columns to scan for an applicability/marks
    note — these often sit in a header-less column just past 'Type Name', so we
    read them positionally rather than by header.
    """
    chapters = []
    cur = None

    def note_of(row):
        for j in range(note_range[0], min(note_range[1], len(row))):
            v = _norm(row[j])
            if v:
                return v
        return ""

    for row in rows[hdr_idx + 1:]:
        get = lambda j: _norm(row[j]) if (j is not None and j < len(row)) else ""
        cname = get(cols["cnm"])
        cno = get(cols["cn"])
        tname = get(cols["tnm"])
        tno = get(cols["tn"])
        note = note_of(row)
        if not (cname or cno or tname):
            continue  # blank spacer row
        if cname:  # new chapter
            cur = {"no": cno.replace(".0", ""), "name": cname,
                   "note": note if not tname else "", "topics": []}
            if super_label:
                cur["group"] = super_label
            chapters.append(cur)
        if tname:
            if cur is None:  # topic with no chapter yet — synthesize
                cur = {"no": cno.replace(".0", ""), "name": cname or "(unspecified)",
                       "note": "", "topics": []}
                if super_label:
                    cur["group"] = super_label
                chapters.append(cur)
            cur["topics"].append({"no": tno.replace("Type No", "").strip() or tno,
                                  "name": tname, "note": note})
    return chapters


def parse(path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    out = {"source": str(path), "subjects": {}}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        if not rows:
            continue
        hdr = _find_header_row(rows)
        if hdr is None:
            continue
        header_cells = [_norm(c) for c in rows[hdr]]
        # super-header row directly above (for multi-table sheets)
        super_row = [_norm(c) for c in rows[hdr - 1]] if hdr > 0 else []
        segs = _segments(header_cells)
        max_col = max((len(r) for r in rows), default=0)
        chapters = []
        for si, seg in enumerate(segs):
            cols = _block_cols(header_cells, seg)
            if not cols:
                continue
            # note column(s): from just past this block's rightmost field up to
            # the next block's first column (or end of row). Catches header-less
            # applicability flags like "Not in GATE".
            known = [x for x in cols.values() if x is not None]
            note_start = max(known) + 1
            note_end = segs[si + 1][0] if si + 1 < len(segs) else max_col
            super_label = ""
            if super_row:
                labels = [super_row[j] for j in seg if j < len(super_row) and super_row[j]]
                super_label = labels[0] if labels else ""
            chapters.extend(
                _parse_block(rows, hdr, cols, (note_start, note_end), super_label))
        if chapters:
            out["subjects"][sheet.strip()] = {"chapters": chapters}
    wb.close()
    return out


def flat_lines(tax) -> list:
    lines = []
    for subj, data in tax["subjects"].items():
        for ch in data["chapters"]:
            if ch["topics"]:
                for t in ch["topics"]:
                    suffix = f"  [{t['note']}]" if t.get("note") else ""
                    lines.append(f"{subj} > {ch['name']} > {t['name']}{suffix}")
            else:
                lines.append(f"{subj} > {ch['name']}")
    return lines


def main(argv):
    import argparse
    ap = argparse.ArgumentParser(description="Parse the ExamDost chapter/topic master sheet.")
    ap.add_argument("path", nargs="?", default=str(DEFAULT),
                    help="Path to the master .xlsx (default: bundled assets copy).")
    ap.add_argument("--flat", action="store_true",
                    help="Print canonical 'Subject > Chapter > Topic' lines instead of JSON.")
    ap.add_argument("--subject", default=None,
                    help="Filter to subjects whose name contains this string.")
    ns = ap.parse_args(argv)
    tax = parse(ns.path)
    if ns.subject:
        tax["subjects"] = {k: v for k, v in tax["subjects"].items()
                           if ns.subject.lower() in k.lower()}
    if ns.flat:
        print("\n".join(flat_lines(tax)))
    else:
        print(json.dumps(tax, ensure_ascii=False, indent=2))
    # summary to stderr so it doesn't pollute piped JSON
    n_ch = sum(len(v["chapters"]) for v in tax["subjects"].values())
    n_t = sum(len(c["topics"]) for v in tax["subjects"].values() for c in v["chapters"])
    print(f"[loaded {len(tax['subjects'])} subjects, {n_ch} chapters, {n_t} topics]",
          file=sys.stderr)


if __name__ == "__main__":
    main(sys.argv[1:])
